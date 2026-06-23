#!/usr/bin/env python3
"""
Excel VLOOKUP + 图片嵌入图形化工具 (tkinter 版)
根据匹配列将表B的图片链接传递到表A，下载并嵌入图片，支持压缩、分批处理。
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import queue
import logging
import sys
import os
import gc
import io
import atexit
import signal
import tempfile
import shutil
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import requests
from PIL import Image
import xlsxwriter
from openpyxl import load_workbook

# ---------- 全局临时文件管理 ----------
_temp_output_path = None

def cleanup():
    global _temp_output_path
    if _temp_output_path and os.path.exists(_temp_output_path):
        try:
            os.unlink(_temp_output_path)
        except Exception:
            pass

atexit.register(cleanup)

# ---------- 图片下载与压缩 ----------
def download_and_compress(url, max_size, quality, timeout):
    try:
        resp = requests.get(url, timeout=timeout, stream=True)
        resp.raise_for_status()
        img = Image.open(resp.raw)

        if max_size:
            img.thumbnail(max_size, Image.LANCZOS)

        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=quality)
        buf.seek(0)
        return buf
    except Exception as e:
        logging.warning(f"下载/压缩失败 {url}: {e}")
        return None

# ---------- 批处理写入 ----------
def process_batch(worksheet, start_row, row_data_list, urls, image_col_idx,
                  executor, max_size, quality, timeout, row_height):
    futures = []
    for url in urls:
        if url and isinstance(url, str) and url.strip():
            fut = executor.submit(download_and_compress, url.strip(), max_size, quality, timeout)
        else:
            fut = None
        futures.append(fut)

    for i, (row_data, fut) in enumerate(zip(row_data_list, futures)):
        current_row = start_row + i
        worksheet.set_row(current_row, row_height)
        worksheet.write_row(current_row, 0, row_data)

        if fut is not None:
            img_bytes = fut.result()
            if img_bytes:
                try:
                    worksheet.insert_image(
                        current_row, image_col_idx,
                        'image.jpg', {'image_data': img_bytes}
                    )
                except Exception as e:
                    logging.warning(f"第 {current_row+1} 行插入图片失败: {e}")
    gc.collect()

# ---------- 核心处理逻辑（在后台线程调用） ----------
def process_task(file_a, sheet_a, col_a,
                 file_b, sheet_b, col_b, col_b_trans,
                 col_image, output_path, in_place,
                 max_size, quality, batch_size, workers, timeout, row_height,
                 log_queue):
    """
    执行主要处理流程，并通过 log_queue 发送消息：
    ('log', level, msg)   日志消息
    ('progress', current, total)  进度更新
    ('done', success_count, fail_count)  完成
    """
    # 自定义日志处理器，将消息放入队列
    class QueueHandler(logging.Handler):
        def __init__(self, queue):
            super().__init__()
            self.queue = queue
        def emit(self, record):
            self.queue.put(('log', record.levelname, self.format(record)))

    # 配置日志
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    handler = QueueHandler(log_queue)
    handler.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
    logger.addHandler(handler)

    try:
        # 输出路径逻辑
        if in_place:
            final_output = file_a
        else:
            final_output = output_path

        # 读取表B
        log_queue.put(('log', 'INFO', '正在读取表B...'))
        try:
            sheet_b_name = int(sheet_b)
        except ValueError:
            sheet_b_name = sheet_b
        df_b = pd.read_excel(file_b, sheet_name=sheet_b_name)

        # 解析列名/索引
        col_b_name = col_b
        col_b_trans_name = col_b_trans
        try:
            col_b_idx = int(col_b)
            col_b_name = df_b.columns[col_b_idx]
        except ValueError:
            pass
        try:
            col_b_trans_idx = int(col_b_trans)
            col_b_trans_name = df_b.columns[col_b_trans_idx]
        except ValueError:
            pass

        if col_b_name not in df_b.columns or col_b_trans_name not in df_b.columns:
            log_queue.put(('log', 'ERROR', f'表B中未找到列: {col_b_name} 或 {col_b_trans_name}'))
            log_queue.put(('done', 0, 0))
            return

        df_b = df_b[[col_b_name, col_b_trans_name]].dropna(subset=[col_b_name])
        lookup = dict(zip(df_b[col_b_name], df_b[col_b_trans_name]))
        del df_b
        gc.collect()

        # 打开表A
        log_queue.put(('log', 'INFO', '正在处理表A...'))
        wb_a = load_workbook(file_a, read_only=True)
        try:
            sheet_a_name = int(sheet_a)
            sheet_a_obj = wb_a.worksheets[sheet_a_name]
        except ValueError:
            sheet_a_obj = wb_a[sheet_a]

        # 读表头
        rows_iter = sheet_a_obj.iter_rows(min_row=1, max_row=1, values_only=True)
        header = next(rows_iter)
        num_orig_cols = len(header)

        # 定位匹配列
        col_a_name = col_a
        try:
            col_a_idx = int(col_a)
            if col_a_idx < 0 or col_a_idx >= num_orig_cols:
                raise IndexError(f'列索引越界: {col_a_idx}')
        except ValueError:
            if col_a not in header:
                log_queue.put(('log', 'ERROR', f'表A中未找到列名: {col_a}'))
                log_queue.put(('done', 0, 0))
                wb_a.close()
                return
            col_a_idx = header.index(col_a)

        # 图片列处理
        if col_image in header:
            image_col_idx = header.index(col_image)
            output_header = list(header)
            log_queue.put(('log', 'INFO', f"列 '{col_image}' 已存在，将覆盖"))
        else:
            image_col_idx = num_orig_cols
            output_header = list(header) + [col_image]
            log_queue.put(('log', 'INFO', f"将在末尾新增列 '{col_image}'"))

        # 计算总行数
        total_rows = sum(1 for _ in sheet_a_obj.iter_rows(min_row=2, values_only=True))
        wb_a.close()
        wb_a = load_workbook(file_a, read_only=True)
        try:
            sheet_a_obj = wb_a.worksheets[int(sheet_a)] if sheet_a.isdigit() else wb_a[sheet_a]
        except:
            pass  # 重新获取

        # 创建临时输出文件
        global _temp_output_path
        fd, _temp_output_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)

        workbook = xlsxwriter.Workbook(_temp_output_path, {'constant_memory': True})
        worksheet = workbook.add_worksheet()

        # 写表头
        for col_idx, val in enumerate(output_header):
            worksheet.write(0, col_idx, val)
        worksheet.set_column(image_col_idx, image_col_idx, 20)

        executor = ThreadPoolExecutor(max_workers=workers)
        success_count = 0
        fail_count = 0

        try:
            row_idx = 1
            batch_rows = []
            batch_urls = []
            processed = 0

            for row in sheet_a_obj.iter_rows(min_row=2, values_only=True):
                match_val = row[col_a_idx]
                url = lookup.get(match_val, None)

                if image_col_idx < num_orig_cols:
                    row_data = list(row)
                    row_data[image_col_idx] = ''
                else:
                    row_data = list(row) + ['']
                batch_rows.append(row_data)
                batch_urls.append(url)

                if len(batch_rows) >= batch_size:
                    process_batch(worksheet, row_idx, batch_rows, batch_urls,
                                  image_col_idx, executor, max_size, quality, timeout, row_height)
                    success_count += sum(1 for u in batch_urls if u)
                    fail_count += sum(1 for u in batch_urls if not u)
                    processed += len(batch_rows)
                    log_queue.put(('progress', processed, total_rows))
                    row_idx += len(batch_rows)
                    batch_rows.clear()
                    batch_urls.clear()
                    gc.collect()

            if batch_rows:
                process_batch(worksheet, row_idx, batch_rows, batch_urls,
                              image_col_idx, executor, max_size, quality, timeout, row_height)
                success_count += sum(1 for u in batch_urls if u)
                fail_count += sum(1 for u in batch_urls if not u)
                processed += len(batch_rows)
                log_queue.put(('progress', processed, total_rows))
        finally:
            executor.shutdown(wait=True)
            workbook.close()
            wb_a.close()

        # 替换输出文件
        shutil.move(_temp_output_path, final_output)
        _temp_output_path = None
        log_queue.put(('log', 'INFO', f'处理完成！结果保存至 {final_output}'))
        log_queue.put(('done', success_count, fail_count))

    except Exception as e:
        log_queue.put(('log', 'ERROR', f'处理异常: {e}'))
        log_queue.put(('done', 0, 0))
    finally:
        logger.removeHandler(handler)

# ---------- GUI 应用 ----------
class App:
    def __init__(self, root):
        self.root = root
        root.title("Excel 图片嵌入工具 (VLOOKUP + 图片下载)")
        root.geometry("650x750")
        root.resizable(True, True)

        # 变量绑定
        self.file_a_var = tk.StringVar()
        self.sheet_a_var = tk.StringVar(value='0')
        self.col_a_var = tk.StringVar()
        self.file_b_var = tk.StringVar()
        self.sheet_b_var = tk.StringVar(value='0')
        self.col_b_var = tk.StringVar()
        self.col_b_trans_var = tk.StringVar()
        self.col_image_var = tk.StringVar(value='Image')
        self.output_var = tk.StringVar(value='output.xlsx')
        self.in_place_var = tk.BooleanVar()
        self.max_width_var = tk.IntVar(value=800)
        self.max_height_var = tk.IntVar(value=800)
        self.quality_var = tk.IntVar(value=70)
        self.batch_size_var = tk.IntVar(value=50)
        self.workers_var = tk.IntVar(value=10)
        self.timeout_var = tk.IntVar(value=30)
        self.row_height_var = tk.IntVar(value=80)
        self.no_compress_var = tk.BooleanVar()

        # 处理线程控制
        self.processing = False
        self.thread = None
        self.queue = queue.Queue()

        self.create_widgets()
        self.periodic_update()

    def create_widgets(self):
        # 文件选择区域
        frame = ttk.LabelFrame(self.root, text="输入文件", padding=10)
        frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(frame, text="表A (需要图片的表格):").grid(row=0, column=0, sticky='w')
        ttk.Entry(frame, textvariable=self.file_a_var, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="浏览", command=lambda: self.browse_file(self.file_a_var)).grid(row=0, column=2)

        ttk.Label(frame, text="表A工作表:").grid(row=1, column=0, sticky='w', pady=2)
        ttk.Entry(frame, textvariable=self.sheet_a_var, width=20).grid(row=1, column=1, sticky='w')

        ttk.Label(frame, text="表A匹配列 (列名/索引):").grid(row=2, column=0, sticky='w', pady=2)
        ttk.Entry(frame, textvariable=self.col_a_var, width=20).grid(row=2, column=1, sticky='w')

        ttk.Separator(frame, orient='horizontal').grid(row=3, columnspan=3, sticky='ew', pady=5)

        ttk.Label(frame, text="表B (包含图片链接):").grid(row=4, column=0, sticky='w')
        ttk.Entry(frame, textvariable=self.file_b_var, width=50).grid(row=4, column=1, padx=5)
        ttk.Button(frame, text="浏览", command=lambda: self.browse_file(self.file_b_var)).grid(row=4, column=2)

        ttk.Label(frame, text="表B工作表:").grid(row=5, column=0, sticky='w', pady=2)
        ttk.Entry(frame, textvariable=self.sheet_b_var, width=20).grid(row=5, column=1, sticky='w')

        ttk.Label(frame, text="表B匹配列 (列名/索引):").grid(row=6, column=0, sticky='w', pady=2)
        ttk.Entry(frame, textvariable=self.col_b_var, width=20).grid(row=6, column=1, sticky='w')

        ttk.Label(frame, text="表B图片链接列 (列名/索引):").grid(row=7, column=0, sticky='w', pady=2)
        ttk.Entry(frame, textvariable=self.col_b_trans_var, width=20).grid(row=7, column=1, sticky='w')

        # 图片设置
        frame2 = ttk.LabelFrame(self.root, text="图片与输出设置", padding=10)
        frame2.pack(fill='x', padx=10, pady=5)

        ttk.Label(frame2, text="图片存放列名:").grid(row=0, column=0, sticky='w')
        ttk.Entry(frame2, textvariable=self.col_image_var, width=15).grid(row=0, column=1, sticky='w')

        ttk.Checkbutton(frame2, text="不压缩图片", variable=self.no_compress_var).grid(row=0, column=2, padx=10)

        ttk.Label(frame2, text="最大宽度:").grid(row=1, column=0, sticky='w', pady=2)
        ttk.Entry(frame2, textvariable=self.max_width_var, width=6).grid(row=1, column=1, sticky='w')
        ttk.Label(frame2, text="px").grid(row=1, column=1, sticky='e')

        ttk.Label(frame2, text="最大高度:").grid(row=1, column=2, sticky='w', pady=2)
        ttk.Entry(frame2, textvariable=self.max_height_var, width=6).grid(row=1, column=3, sticky='w')
        ttk.Label(frame2, text="px").grid(row=1, column=3, sticky='e')

        ttk.Label(frame2, text="JPEG质量 (1-100):").grid(row=2, column=0, sticky='w', pady=2)
        ttk.Entry(frame2, textvariable=self.quality_var, width=6).grid(row=2, column=1, sticky='w')

        ttk.Label(frame2, text="批处理行数:").grid(row=2, column=2, sticky='w', pady=2)
        ttk.Entry(frame2, textvariable=self.batch_size_var, width=6).grid(row=2, column=3, sticky='w')

        ttk.Label(frame2, text="线程数:").grid(row=3, column=0, sticky='w', pady=2)
        ttk.Entry(frame2, textvariable=self.workers_var, width=6).grid(row=3, column=1, sticky='w')

        ttk.Label(frame2, text="下载超时(秒):").grid(row=3, column=2, sticky='w', pady=2)
        ttk.Entry(frame2, textvariable=self.timeout_var, width=6).grid(row=3, column=3, sticky='w')

        ttk.Label(frame2, text="行高(磅):").grid(row=4, column=0, sticky='w', pady=2)
        ttk.Entry(frame2, textvariable=self.row_height_var, width=6).grid(row=4, column=1, sticky='w')

        # 输出选项
        frame3 = ttk.LabelFrame(self.root, text="输出", padding=10)
        frame3.pack(fill='x', padx=10, pady=5)

        ttk.Label(frame3, text="输出文件:").grid(row=0, column=0, sticky='w')
        ttk.Entry(frame3, textvariable=self.output_var, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(frame3, text="浏览", command=self.browse_save).grid(row=0, column=2)
        ttk.Checkbutton(frame3, text="直接覆盖原表A (in-place)", variable=self.in_place_var,
                        command=self.toggle_in_place).grid(row=1, column=0, columnspan=2, sticky='w')

        # 控制按钮
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(fill='x', padx=10, pady=10)
        self.start_btn = ttk.Button(btn_frame, text="开始处理", command=self.start_processing)
        self.start_btn.pack(side='left', padx=5)
        ttk.Button(btn_frame, text="退出", command=self.root.quit).pack(side='right', padx=5)

        # 进度条
        self.progress = ttk.Progressbar(self.root, orient='horizontal', length=500, mode='determinate')
        self.progress.pack(padx=10, pady=5, fill='x')

        # 日志显示
        log_frame = ttk.LabelFrame(self.root, text="运行日志", padding=5)
        log_frame.pack(fill='both', expand=True, padx=10, pady=5)
        self.log_text = tk.Text(log_frame, height=10, wrap='word', state='disabled')
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

    def browse_file(self, var):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if filename:
            var.set(filename)

    def browse_save(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filename:
            self.output_var.set(filename)

    def toggle_in_place(self):
        if self.in_place_var.get():
            self.output_var.set('')  # 不需要输出路径
            messagebox.showinfo("提示", "已启用原地修改，将直接覆盖表A文件。请确保文件已关闭。")

    def start_processing(self):
        if self.processing:
            messagebox.showwarning("警告", "正在处理中，请稍后...")
            return

        # 基本输入检查
        if not self.file_a_var.get() or not self.file_b_var.get():
            messagebox.showerror("错误", "请选择表A和表B文件。")
            return
        if not self.col_a_var.get() or not self.col_b_var.get() or not self.col_b_trans_var.get():
            messagebox.showerror("错误", "请填写所有匹配列。")
            return
        if not self.in_place_var.get() and not self.output_var.get():
            messagebox.showerror("错误", "请指定输出文件路径或勾选原地修改。")
            return

        self.processing = True
        self.start_btn.config(state='disabled')
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        self.progress['value'] = 0

        # 收集参数
        params = {
            'file_a': self.file_a_var.get(),
            'sheet_a': self.sheet_a_var.get(),
            'col_a': self.col_a_var.get(),
            'file_b': self.file_b_var.get(),
            'sheet_b': self.sheet_b_var.get(),
            'col_b': self.col_b_var.get(),
            'col_b_trans': self.col_b_trans_var.get(),
            'col_image': self.col_image_var.get(),
            'output_path': self.output_var.get(),
            'in_place': self.in_place_var.get(),
            'max_size': None if self.no_compress_var.get() else (self.max_width_var.get(), self.max_height_var.get()),
            'quality': 95 if self.no_compress_var.get() else self.quality_var.get(),
            'batch_size': self.batch_size_var.get(),
            'workers': self.workers_var.get(),
            'timeout': self.timeout_var.get(),
            'row_height': self.row_height_var.get(),
            'log_queue': self.queue
        }

        self.thread = threading.Thread(target=process_task, kwargs=params, daemon=True)
        self.thread.start()

    def periodic_update(self):
        """定时检查队列，更新界面"""
        try:
            while True:
                msg = self.queue.get_nowait()
                if msg[0] == 'log':
                    _, level, text = msg
                    self.append_log(f"[{level}] {text}")
                elif msg[0] == 'progress':
                    _, current, total = msg
                    if total > 0:
                        self.progress['value'] = (current / total) * 100
                    self.root.update_idletasks()
                elif msg[0] == 'done':
                    _, success, fail = msg
                    self.append_log(f"完成：图片下载成功 {success}，失败/空链接 {fail}")
                    self.processing = False
                    self.start_btn.config(state='normal')
                    messagebox.showinfo("完成", f"处理完成！\n成功: {success}\n失败/空链接: {fail}")
        except queue.Empty:
            pass
        self.root.after(100, self.periodic_update)

    def append_log(self, text):
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, text + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')

# ---------- 主入口 ----------
if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()